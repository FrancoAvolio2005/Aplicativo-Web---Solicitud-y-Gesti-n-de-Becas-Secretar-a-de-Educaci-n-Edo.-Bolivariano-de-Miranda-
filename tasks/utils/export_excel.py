import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import datetime

# Importaciones de Django para la obtención de datos y modelos
from django.db.models import Count, Q 
from django.contrib.auth.models import User 

# Importaciones de Modelos
from ..models import Solicitud, Profile, Becas, Plantel 

# =============================
# 1. ESTILOS Y MAPPINGS COMUNES 
# =============================

HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

# Mapeos de CHOICES
ESTADO_CHOICES = (
    ('A', 'Amazonas'), ('AN', 'Anzoátegui'), ('AP', 'Apure'), ('AR', 'Aragua'), 
    ('B', 'Barinas'), ('BO', 'Bolívar'), ('CA', 'Carabobo'), ('C', 'Cojedes'),
)
DEPENDENCIA_CHOICES = (
    ('NAC', 'Nacional'), ('EST', 'Estadal'), ('MUN', 'Municipal'),
)
MODALIDAD_CHOICES = (
    ('P', 'Presencial'), ('S', 'Semipresencial'), ('D', 'Distancia'),
)
ESTATUS_CHOICES_PLANTEL = (
    ('A', 'Activo'), ('I', 'Inactivo'),
)

estado_mapping = dict(ESTADO_CHOICES)
dependencia_mapping = dict(DEPENDENCIA_CHOICES)
modalidad_mapping = dict(MODALIDAD_CHOICES)
estatus_plantel_mapping = dict(ESTATUS_CHOICES_PLANTEL)


def _apply_header_style(sheet, headers, fill_color):
    """Función helper para aplicar estilos comunes a la fila de encabezado."""
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    for col_num, header_text in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = header_fill
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        sheet.column_dimensions[get_column_letter(col_num)].width = 25

def _apply_row_style(sheet, row_num, col_count):
    """Función helper para aplicar estilos comunes a las filas de datos."""
    for col_num in range(1, col_count + 1):
        cell = sheet.cell(row=row_num, column=col_num)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)


# ==========================================
# 2. INTERFAZ DE ESTRATEGIA (ExportStrategy)
# ==========================================

class ExportStrategy:
    """Clase base (Interfaz) para todas las estrategias de exportación."""
    def export(self):
        """Método que debe ser implementado por las estrategias concretas."""
        raise NotImplementedError("La subclase debe implementar el método export()")

    def get_headers(self):
        """Define los encabezados específicos de la exportación."""
        raise NotImplementedError("La subclase debe definir los encabezados.")

    def get_data(self):
        """Consulta y prepara los datos del modelo."""
        raise NotImplementedError("La subclase debe implementar la lógica de consulta.")

    def get_title(self):
        """Define el título de la hoja."""
        return "Reporte"


# ========================
# 3. ESTRATEGIAS CONCRETAS
# ========================

class ProfilesExportStrategy(ExportStrategy):
    """Estrategia para exportar perfiles."""
    def get_title(self):
        return "Reporte de Solicitantes"
    
    def get_headers(self):
        return [
            "ID de Usuario", "Nombre de Usuario", "Correo Electrónico",
            "Nombre Completo", "Apellido Completo", "Cédula de Identidad",
            "Edad", "Género", "Fecha de Nacimiento", "Número de Teléfono"
        ]
        
    def get_data(self):
        return Profile.objects.filter(is_analista_exterior=False).select_related('user').exclude(user__is_superuser=True)
    
    def export(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = self.get_title()
        
        headers = self.get_headers()
        _apply_header_style(sheet, headers, fill_color="4CAF50")
        
        # Ajustes de ancho específicos
        sheet.column_dimensions[get_column_letter(4)].width = 35 
        sheet.column_dimensions[get_column_letter(5)].width = 35 
        sheet.column_dimensions[get_column_letter(3)].width = 35 

        for profile in self.get_data():
            row_data = [
                profile.user.id,
                profile.user.username,
                profile.user.email,
                profile.nombre_completo,
                profile.apellido_completo,
                profile.cedula_identidad,
                profile.edad,
                profile.genero,
                profile.fecha_nacimiento.strftime('%d-%m-%Y') if profile.fecha_nacimiento else '',
                profile.numero_telefono
            ]
            sheet.append(row_data)
            _apply_row_style(sheet, sheet.max_row, len(headers))
            
        return workbook


class BecasExportStrategy(ExportStrategy):
    """Estrategia para exportar el catálogo de becas."""
    def get_title(self):
        return "Reporte de Becas"
        
    def get_headers(self):
        return ["ID Beca", "Nombre", "Descripción"]
        
    def get_data(self):
        return Becas.objects.all()
        
    def export(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = self.get_title()
        
        headers = self.get_headers()
        _apply_header_style(sheet, headers, fill_color="28A745")
        
        sheet.column_dimensions[get_column_letter(3)].width = 80 

        for beca in self.get_data():
            row_data = [
                beca.id_beca,
                beca.nombre,
                beca.descripcion
            ]
            sheet.append(row_data)
            _apply_row_style(sheet, sheet.max_row, len(headers))

        return workbook


class PlantelesExportStrategy(ExportStrategy):
    """Estrategia para exportar el catálogo de planteles."""
    def get_title(self):
        return "Reporte de Planteles"

    def get_headers(self):
        return [
            "Nombre del Plantel", "Estado del Plantel", "Municipio del Plantel",
            "Código del Plantel", "Tipo de Dependencia", "Modalidad Principal",
            "Estatus del Plantel"
        ]
        
    def get_data(self):
        return Plantel.objects.all()
        
    def export(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = self.get_title()
        
        headers = self.get_headers()
        _apply_header_style(sheet, headers, fill_color="007BFF")
        
        for col_num in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(col_num)].width = 30 

        for plantel in self.get_data():
            row_data = [
                plantel.nombre_plantel,
                estado_mapping.get(getattr(plantel, 'estado_plantel', ''), getattr(plantel, 'estado_plantel', '')),
                getattr(plantel, 'municipio_plantel', ''),
                getattr(plantel, 'codigo_plantel', ''),
                dependencia_mapping.get(getattr(plantel, 'tipo_dependencia', ''), getattr(plantel, 'tipo_dependencia', '')),
                modalidad_mapping.get(getattr(plantel, 'modalidad_principal', ''), getattr(plantel, 'modalidad_principal', '')),
                estatus_plantel_mapping.get(getattr(plantel, 'estatus_plantel', ''), getattr(plantel, 'estatus_plantel', ''))
            ]
            sheet.append(row_data)
            _apply_row_style(sheet, sheet.max_row, len(headers))
            
        return workbook


class SolicitudesExportStrategy(ExportStrategy):
    """Estrategia para exportar el reporte completo de solicitudes."""
    def get_title(self):
        return "Reporte de Solicitudes"
        
    def get_headers(self):
        return [
            "Nombre del Solicitante", "Apellido del Solicitante",
            "Nombre del estudiante", "Apellido del estudiante",
            "Cédula", "Teléfono", "Beca Solicitada", "Estatus"
        ]
        
    def get_data(self):
        estatus_orden = ['Asignada', 'Aprobada', 'Rechazada', 'En proceso']
        solicitudes_ordenadas = []

        for estatus_nombre in estatus_orden:
            solicitudes = Solicitud.objects.filter(
                estatus_beca__nombre=estatus_nombre,
                user__isnull=False
            ).select_related('user__profile', 'beca', 'estatus_beca')
            solicitudes_ordenadas.extend(list(solicitudes))
        return solicitudes_ordenadas
        
    def export(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.get_title()
        
        headers = self.get_headers()
        _apply_header_style(ws, headers, fill_color="4CAF50")
        
        ws.column_dimensions[get_column_letter(3)].width = 25
        ws.column_dimensions[get_column_letter(4)].width = 25

        for solicitud in self.get_data():
            profile = solicitud.user.profile if hasattr(solicitud.user, 'profile') else None
            
            nombre_completo = getattr(profile, 'nombre_completo', "")
            apellido_completo = getattr(profile, 'apellido_completo', "")
            cedula_identidad = getattr(profile, 'cedula_identidad', "")
            numero_telefono = getattr(profile, 'numero_telefono', "")
                
            beca_nombre = solicitud.beca.nombre if solicitud.beca else ""
            estatus_actual = solicitud.estatus_beca.nombre if solicitud.estatus_beca else ""
            
            nombre_becario = getattr(solicitud, 'nombre_becario', "")
            apellido_becario = getattr(solicitud, 'apellido_becario', "")

            row_data = [
                nombre_completo,
                apellido_completo,
                nombre_becario,
                apellido_becario,
                cedula_identidad,
                numero_telefono,
                beca_nombre,
                estatus_actual
            ]
            ws.append(row_data)
            _apply_row_style(ws, ws.max_row, len(row_data))

        return wb


# ================================
# 4. CONTEXTO (ExcelExportContext) y MAPEO
# ================================

class ExcelExportContext:
    """
    El Contexto que utiliza una Estrategia concreta para ejecutar la exportación.
    """
    def __init__(self, strategy: ExportStrategy):
        self._strategy = strategy

    def execute_export(self):
        """Ejecuta el método de exportación de la estrategia seleccionada."""
        return self._strategy.export()

# Mapa que relaciona el tipo de reporte con la Estrategia concreta a usar
STRATEGY_MAP = {
    'profiles': ProfilesExportStrategy,
    'becas': BecasExportStrategy,
    'planteles': PlantelesExportStrategy,
    'solicitudes': SolicitudesExportStrategy,
}


# ================================================
# 5. PUNTO DE ENTRADA (Lo que las vistas llamarán)
# ================================================

def get_exporter(report_type: str):
    """
    Función de fábrica que devuelve el Contexto de Exportación para un tipo de reporte.
    
    Ejemplo de uso en la vista:
    workbook = get_exporter('profiles').execute_export()
    """
    strategy_class = STRATEGY_MAP.get(report_type)
    
    if not strategy_class:
        raise ValueError(f"Estrategia de exportación desconocida: {report_type}")
        
    strategy_instance = strategy_class()
    return ExcelExportContext(strategy_instance)